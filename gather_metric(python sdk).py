#!/usr/bin/env python3

from openpyxl import load_workbook
import os
import subprocess
from azure.cli.core import get_default_cli
from azhelper import *
# az login --service-principal -u http://sample-cli-login -p Test1234 --tenant 54826b22-38d6-4fb2-bad9-b7b93a3e9c5a


def az_cli(args_str):
    args = args_str.split()
    cli = get_default_cli()
    cli.invoke(args)
    if cli.result.result:
        return cli.result.result
    elif cli.result.error:
        raise cli.result.error
    return True


wb = load_workbook("/Users/hojaelee/Desktop/powershell/gather_metric.xlsx")
ws = wb['Metrics']

os.system('az account set -s ca9621d4-bfba-4a66-8f48-c4c7eaa0a21f')

# ws.cell(row=3, column=3, value=55)
# ws["C3"] = 3
# print(ws["A2"].value)

# get vm's resource id
vm_id = os.popen(
    'az vm show -g '+str(ws.cell(2, 1).value)+' -n '+str(ws.cell(2, 2).value)+' --query id').read()

#test = get_default_cli().invoke(['vm', 'list', '-g', 'linux-stress_group'])
# print(test)

response = az_cli("vm list")
print("vm's: %s" % (response))

# raw_data = os.popen(('az monitor metrics list --resource '+vm_id[:-1] +
#                     ' --metric "Percentage CPU" --interval 1m --offset 30d --aggregation Average | grep average')).read()


# print(raw_data)
wb.save("/Users/hojaelee/desktop/powershell/gather_metric.xlsx")
