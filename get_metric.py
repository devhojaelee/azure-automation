from openpyxl import load_workbook
import os
import subprocess

wb = load_workbook("/Users/hojaelee/Desktop/powershell/gather_metric.xlsx")
ws = wb['Metrics']

os.system('az account set -s ca9621d4-bfba-4a66-8f48-c4c7eaa0a21f')

vm_id = os.popen(
    'az vm show -g '+str(ws.cell(2, 1).value)+' -n '+str(ws.cell(2, 2).value)+' --query id').read()

raw_data = os.popen(('az monitor metrics list --resource '+vm_id[:-1] +
                     ' --metric "Percentage CPU" --interval 1m --offset 30d --aggregation Average | grep average')).read()
raw_data = raw_data()
print(raw_data)

wb.save("/Users/hojaelee/desktop/powershell/gather_metric.xlsx")
