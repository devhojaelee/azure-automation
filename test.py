from openpyxl import load_workbook, Workbook


# data_only=Ture로 해줘야 수식이 아닌 값으로 받아온다.
load_wb = load_workbook(
    "/Users/hojaelee/desktop/powershell/gather_metric.xlsx", data_only=True)

# 시트 이름으로 불러오기
load_ws = load_wb['Metrics']


#셀 주소에 갑 입력
load_ws['C5'] = 6666

# 셀 주소로 값 출력
print(load_ws['A1'].value)

# 셀 좌표로 값 출력
print(load_ws.cell(1, 2).value)


load_wb.save("/Users/hojaelee/desktop/powershell/gather_metric.xlsx")
