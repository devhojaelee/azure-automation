# from sla import *
from xml.etree.ElementTree import register_namespace
from openpyxl import load_workbook
import os
import subprocess
import warnings
import json  # for string to json
warnings.simplefilter("ignore")

wb = load_workbook("/Users/hojaelee/Desktop/powershell/hanwha.xlsx")
ws = wb['3.3 SLA 분석 평가']
os.system('az account set -s ca9621d4-bfba-4a66-8f48-c4c7eaa0a21f')

# type이 Virtual machine, SQL Virtual machine일 경우, SLA 계산
# 우선 VM이 SLA 계산도 고려해야 할 것도 많고 복잡해서 자동화 타겟으로 정함.

# 1. 동일한 Region에서 / 2개 이상의 AZ에 / 배포된 2개 이상의 인스턴스는 /, 그 중 적어도 한 인스턴스는 99.99% /
# 그럼 가장 높은 SLA의 VM만 99.99% /, 나머지는 single instance SLA 따라감
# 2. 동일한 AVset / or Dedicated host group에서 / 배포된 2개 이상의 인스턴스는 /, 그 중 적어도 한 인스턴스는 99.95% /
# 그럼 가장 높은 SLA의 VM만 99.99% /, 나머지는 single instance SLA 따라감

# Future work : 1) 엑셀에서 SLA 고려안 할 대상 행 삭제

max_row = 2000  # (수동으로 최대 row 값 입력해줄 것. ws.max_row 이 값을 제대로 못읽어옴.)
current_row = 6  # 엑셀에서 데이터가 6행부터 시작
# API 호출 한번만 하기 위해서 vm_data_obj 객체 생성.
vm_data_obj = json.loads(os.popen(('az vm list --output json')).read())
vm_cnt = 0  # vm 넘버링. vm_data_obj[0], vm_data_obj[1]... 각각이 VM 한 대.
vm_list = {}  # VM 이름 뒤에 -1, -2 등을 제거했을 때, 같은 이름의 VM이 있는지 확인하기 위해서.
stripped_vm_name = []
azset_cnt = 0
while max_row > 0:
    if ws['D'+str(current_row)].value == ('Virtual machine' or 'SQL virtual machine'):
        rg_name = str(ws['E'+str(current_row)].value)  # 엑셀에서 rg, vm name 읽어옴
        vm_name = str(ws['C'+str(current_row)].value)
        region = str(ws['F'+str(current_row)].value)

        print("vm name : "+vm_data_obj[vm_cnt]["name"])
        print('vm_cnt : '+str(vm_cnt))

        # 한 VM을 여러 AZ에 배포했을 때, 각 AZ에 name-1, name-2, name-3 으로 배포된다.
        # [:-2]을 해서 같은 name의 VM이 있는지 확인하고, 있으면 [-2]의 값이 1,2,3 인지 확인한다.

        stripped_vm_name = vm_data_obj[vm_cnt]["name"][:-2]
        stripped_num = vm_data_obj[vm_cnt]["name"][-3]
        print("stripped_vm_name : "+stripped_vm_name)
        print("stripped_num : "+stripped_num)

        try:
            vm_list[stripped_vm_name] += 1
        except KeyError:
            vm_list[stripped_vm_name] = 1  # dict[key] 값이 없으면 강제로 넣어줌.

        # 같은 name의 VM이 2개 이상 있으면, 기존 name과 stripped된 name +'-숫자'가 같은지 확인한다.
        # ex. vm_list = {'stress-lin': 1, 'stress-li': 1, 'stress-linux': 3}

        total_data_disks = len(
            vm_data_obj[vm_cnt]["storageProfile"]["dataDisks"])
        print("total_data_disks : "+str(total_data_disks))
        os_disk = []
        data_disk = []
        data_disk_cnt = 0

        # data disk type check
        while total_data_disks > data_disk_cnt:
            data_disk.append(vm_data_obj[vm_cnt]["storageProfile"]["dataDisks"][data_disk_cnt]
                             ["managedDisk"]["storageAccountType"])
            data_disk_cnt += 1
        # os disk type check
        os_disk = vm_data_obj[vm_cnt]["storageProfile"]["osDisk"]["managedDisk"]["storageAccountType"]
        try:  # VM 꺼져있을 때를 대비한 error exception code
            print("os_disk : "+os_disk)
        except TypeError:
            print("VM이 꺼져있습니다.")
        print('disk_data : '+str(data_disk))

        # Single vm instance
        pre_ssd_cnt = 0
        std_ssd_cnt = 0
        std_hdd_cnt = 0

        # os_disk type count
        if os_disk == 'Premium_LRS':
            pre_ssd_cnt += 1
        elif os_disk == 'StandardSSD_LRS':
            std_ssd_cnt += 1
        elif os_disk == 'Standard_LRS':
            std_hdd_cnt += 1
        elif os_disk == None:
            pass
        else:
            print("ERROR. ultradisk를 사용중입니다. 스크립트 개발자에게 반드시 알려주세요.")

        # data_disks type count
        print('<data disk lists>')
        for idx, value in enumerate(data_disk):
            print(str(idx)+' : '+value)
            if value == 'Premium_LRS':
                pre_ssd_cnt += 1
            elif value == 'StandardSSD_LRS':
                std_ssd_cnt += 1
            elif value == 'Standard_LRS':
                std_hdd_cnt += 1
            # value == None을 확인할 필요가 없어. data_disk에 순회할 element가 없으면, for문 곧바로 break 됨.
            else:
                print("ERROR. ultradisk를 사용중입니다. 스크립트 개발자에게 반드시 알려주세요.")

        print('Num of Premium Ssd : '+str(pre_ssd_cnt))
        print('Num of Standard Ssd : '+str(std_ssd_cnt))
        print('Num of Standrad Hdd : '+str(std_hdd_cnt))
        print(vm_list)
        # Single VM instance's SLA check
        if std_hdd_cnt >= 1:
            ws['K'+str(current_row)].value = '99%'
            print("99%\n------\n")
        elif std_hdd_cnt == 0 and std_ssd_cnt >= 1:
            ws['K'+str(current_row)].value = '99.5%'
            print("99.5%\n------\n")
        elif std_hdd_cnt == 0 and std_ssd_cnt == 0 and pre_ssd_cnt >= 1:
            ws['K'+str(current_row)].value = '99.9%'
            print("99.9%\n------\n")
        else:
            # 반드시 os disk는 존재하므로, os_disk가 ultradisk 아닌 이상 위 3개 조건에 해당하지 않으면 vm stopped 간주 가능.
            ws['K'+str(current_row)].value = 'VM stopped'
            print("해당 VM이 켜져있는지 확인하세요.\n------\n")

        vm_cnt += 1

    current_row += 1
    max_row -= 1

# 이거 이렇게 하면 안된다. 즉석에서 바로 비교해야될듯?
print('vm_list = '+str(vm_list))
for k, v in vm_list.items():
    if v >= 2:
        print('key = '+k+' , value = '+str(v))
        for idx in range(v):
            print(vm_data_obj[vm_cnt]["name"])
            print(stripped_vm_name+'-'+str(idx+1))
            if vm_data_obj[vm_cnt]["name"] == stripped_vm_name+'-'+str(idx+1):
                azset_cnt += 1

                print("azset_cnt = "+azset_cnt)

    if azset_cnt == v:
        print(vm_list[stripped_vm_name]+'는 AZset에 존재합니다.')

wb.save("/Users/hojaelee/desktop/powershell/hanwha.xlsx")
